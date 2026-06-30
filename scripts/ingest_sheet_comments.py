#!/usr/bin/env python3
"""Ingest reviewer comments from the shared Google Sheet, map each to its mapping
row, attach to that row's `notes` (idempotent), and write a review doc.

Comments survive in the .xlsx export (xl/comments*.xml). Each is anchored to a cell;
we resolve sheet -> path (col A) -> the mapping CSV row. No auth (link-shared sheet).
Usage: python3 scripts/ingest_sheet_comments.py [--dry-run]
"""
import sys, io, re, urllib.request
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SHEET_ID = "1vqPsLLaV6pMCDFXAReqczompNTnTPEqN"
TABS = {"KDK oncology": "mapping_kdk_oncology.csv",
        "KDK rare diseases": "mapping_kdk_rarediseases.csv",
        "GRZ": "mapping_grz.csv"}
MARK = "CLIN-REVIEW:"

def clean(t):
    lines = [l for l in t.splitlines() if l.strip()]
    lines = [l for l in lines if not l.startswith("======") and not l.strip().startswith("ID#")]
    lines = [l for l in lines if not re.match(r".+\(\d{4}-\d\d-\d\d \d\d:\d\d:\d\d\)\s*$", l)]
    return re.sub(r"\s+", " ", " ".join(lines)).strip()

def fetch_comments():
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    data = urllib.request.urlopen(urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=40).read()
    wb = load_workbook(io.BytesIO(data))
    out = []
    for ws in wb.worksheets:
        if ws.title not in TABS:
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    out.append({"sheet": ws.title, "col": ws.cell(row=1, column=cell.column).value,
                                "path": ws.cell(row=cell.row, column=1).value, "value": cell.value,
                                "author": "Lucien Clin" if "Lucien Clin" in cell.comment.text else (cell.comment.author or "?"),
                                "comment": clean(cell.comment.text)})
    out.sort(key=lambda r: (r["sheet"], r["path"] or ""))
    return out

def main():
    import csv
    dry = "--dry-run" in sys.argv
    cm = fetch_comments()
    print(f"{len(cm)} comments fetched\n")
    # attach to notes per CSV
    by = {}
    for c in cm:
        by.setdefault(TABS[c["sheet"]], {})[c["path"]] = c
    for fn, pmap in by.items():
        rows = list(csv.DictReader(open(ROOT / "mapping" / fn))); fields = list(rows[0].keys()); n = 0
        for r in rows:
            c = pmap.get(r["path"])
            if not c:
                continue
            base = re.sub(r"(\s*\|\s*)?" + re.escape(MARK) + r".*$", "", r["notes"]).strip()  # idempotent
            r["notes"] = (base + " | " if base else "") + f"{MARK} ({c['author']}) {c['comment']}"
            n += 1
        if not dry:
            with open(ROOT / "mapping" / fn, "w", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(rows)
        print(f"  {fn}: {n} rows annotated" + (" [dry]" if dry else ""))
    # review doc
    doc = ["# Reviewer comments from the mapping Google Sheet", "",
           f"Ingested {len(cm)} cell comments (all by Lucien Clin, DNPM model author) and",
           "attached to the matching `notes` in the mapping CSVs. Re-run `python3 scripts/ingest_sheet_comments.py`.", ""]
    for c in cm:
        doc += [f"### [{c['sheet']}] `{c['path']}`",
                f"- cell `{c['col']}` = `{c['value']}`",
                f"- **{c['author']}:** {c['comment']}", ""]
    if not dry:
        (ROOT / "docs" / "SHEET-COMMENTS-clin.md").write_text("\n".join(doc))
        print("\nwrote docs/SHEET-COMMENTS-clin.md")

if __name__ == "__main__":
    main()
