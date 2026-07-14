#!/usr/bin/env python3
"""Export the mapping CSVs (FHIR/MII-KDS + OMOP) to one multi-tab .xlsx for Google Sheets.
Column-generic: each tab uses its own CSV header. Source cols (path..description) shaded
grey/read-only-by-convention; target cols editable; status column gets a dropdown + colours.
Usage: python3 scripts/export_xlsx.py  ->  mapping/mapping-table.xlsx
Round-trip back: in Sheets, File > Download > CSV per tab into mapping/<name>.csv.
"""
import csv
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SHEETS = [  # (tab title, csv) — FHIR/MII-KDS tabs ONLY in the shared sheet.
    # OMOP tables are kept OUT of the Google Sheet by request (they live in git:
    # mapping/mapping_omop_*.csv). Do not add OMOP tabs here.
    ("KDK oncology", "mapping_kdk_oncology.csv"),
    ("KDK rare diseases", "mapping_kdk_rarediseases.csv"),
    ("GRZ", "mapping_grz.csv"),
]
SRC = ["path", "type", "required", "array", "enum", "format", "description"]
STATUS = ["MAPPED", "DRAFT", "NOMAP", "TODO", "REMOVED?"]
FILL = {"MAPPED": "C6EFCE", "DRAFT": "FFEB9C", "NOMAP": "E5E7EB", "TODO": "FCE4D6", "REMOVED?": "F8CBAD"}
WIDE = {"enum": 22, "description": 30, "transform": 40, "omop_transform": 40, "fhir_element": 34,
        "mii_profile": 34, "omop_field": 30, "omop_vocab": 26, "notes": 40, "path": 42, "reference": 40}
hdr_src = PatternFill("solid", fgColor="D9D9D9"); hdr_tgt = PatternFill("solid", fgColor="BDD7EE")
src_fill = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="DDDDDD"); border = Border(thin, thin, thin, thin)

wb = Workbook()
lg = wb.active; lg.title = "README"
lg["A1"] = "genomDE Datenkranz → mapping tables (FHIR/MII-KDS + OMOP CDM)"; lg["A1"].font = Font(bold=True, size=14)
for i, t in enumerate([
 "", "Three FHIR/MII-KDS mapping tabs (GRZ & KDK kept separate). The OMOP-CDM tables are",
 "kept OUT of this sheet by request — they live in git (mapping/mapping_omop_*.csv).",
 "Columns A–G (grey) = schema-derived skeleton: DO NOT EDIT.",
 "Target cols: mii_module, mii_profile, fhir_element, transform, status, notes.",
 "status dropdown: MAPPED / DRAFT / NOMAP / TODO / REMOVED?  (colour-coded).",
 "",
 "Authoritative source of truth = the CSVs in git (KohlbacherLab/genomDE2FHIR, mapping/).",
 "Pull edits back with scripts/sync_from_sheet.py (XLSX export — preserves cell comments).",
 "Reviewer comments flow via native cell COMMENTS (ingest_sheet_comments.py), not the notes column.",
], start=2):
    lg.cell(row=i, column=1, value=t)
lg.column_dimensions["A"].width = 110

for title, fn in SHEETS:
    rows = list(csv.DictReader(open(ROOT / "mapping" / fn)))
    cols = list(rows[0].keys())
    ws = wb.create_sheet(title)
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=name); cell.font = Font(bold=True)
        cell.fill = hdr_src if name in SRC else hdr_tgt; cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = WIDE.get(name, 11)
    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(name, "")); cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=name in WIDE and name != "path")
            if name in SRC:
                cell.fill = src_fill
    last = len(rows) + 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{last}"
    if "status" in cols:
        sc = get_column_letter(cols.index("status") + 1)
        dv = DataValidation(type="list", formula1='"' + ",".join(STATUS) + '"', allow_blank=True)
        ws.add_data_validation(dv); dv.add(f"{sc}2:{sc}{last}")
        for st, hexc in FILL.items():
            ws.conditional_formatting.add(f"{sc}2:{sc}{last}",
                CellIsRule(operator="equal", formula=[f'"{st}"'], fill=PatternFill("solid", fgColor=hexc)))

# ---- computed "Open Issues" tab: every DRAFT/TODO row + rows explicitly flagged for action ----
# word-boundary so "confirm"/"verify" match but resolved prose ("confirmed", "verification") does not
FLAG_RE = re.compile(r"\b(confirm|verify|caveat|unreliable|questionable|tbd)\b", re.I)
oi = wb.create_sheet("Open Issues", index=1)   # right after README
for c, name in enumerate(["tab", "path", "status", "mii_module", "fhir_element", "issue / reason"], start=1):
    cell = oi.cell(row=1, column=c, value=name); cell.font = Font(bold=True); cell.fill = hdr_tgt; cell.border = border
for c, w in enumerate([16, 46, 9, 13, 32, 70], start=1):
    oi.column_dimensions[get_column_letter(c)].width = w
r = 2
for title, fn in SHEETS:
    for row in csv.DictReader(open(ROOT / "mapping" / fn)):
        st = (row.get("status") or "").upper()
        text = (row.get("transform", "") or "") + " " + (row.get("notes", "") or "")
        if st not in ("DRAFT", "TODO") and not FLAG_RE.search(text):
            continue
        seg = [s.strip() for s in (row.get("notes") or "").split("|")
               if any(k in s.lower() for k in ("num-omics", "clin-resolve", "verify", "confirm", "caveat"))]
        reason = seg[-1] if seg else (row.get("transform", "")[:140])
        for c, v in enumerate([title, row["path"], row.get("status", ""), row.get("mii_module", ""),
                               row.get("fhir_element", ""), reason], start=1):
            cell = oi.cell(row=r, column=c, value=v); cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=c in (5, 6))
        hexc = FILL.get(st)
        if hexc:
            oi.cell(row=r, column=3).fill = PatternFill("solid", fgColor=hexc)
        r += 1
oi.freeze_panes = "A2"; oi.auto_filter.ref = f"A1:F{r - 1}"

out = ROOT / "mapping" / "mapping-table.xlsx"; wb.save(out)
print(f"wrote {out.relative_to(ROOT)} ({len(SHEETS)} data tabs + README + Open Issues [{r - 2}])")
