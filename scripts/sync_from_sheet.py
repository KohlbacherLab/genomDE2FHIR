#!/usr/bin/env python3
"""Pull mapping edits from the shared Google Sheet back into the canonical CSVs.

Only the TARGET columns (mii_module, mii_profile, fhir_element, transform, status,
notes) are pulled, matched by `path`. The schema-derived source columns (A-G) are
LEFT UNTOUCHED — they are owned by scripts/regen-mapping.sh, not the sheet.

Pulls via the XLSX export (NOT the CSV/gviz export): CSV export silently drops cell
comments, XLSX preserves them. The sheet must be link-shared ("anyone with the link
can view"). No auth needed. (Comments themselves: scripts/ingest_sheet_comments.py.)

Usage:
  python3 scripts/sync_from_sheet.py [--dry-run] [--id <sheetId>]
"""
import sys, io, csv, urllib.request
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SHEET_ID = "1vqPsLLaV6pMCDFXAReqczompNTnTPEqN"   # KohlbacherLab genomDE2FHIR mapping table
TABS = {
    "KDK oncology": "mapping_kdk_oncology.csv",
    "KDK rare diseases": "mapping_kdk_rarediseases.csv",
    "GRZ": "mapping_grz.csv",
}
# NB: `notes` is intentionally NOT pulled — it is repo-managed (CLIN-REVIEW comment
# annotations come from scripts/ingest_sheet_comments.py + alignment-script provenance).
# Reviewer feedback flows via the sheet's native cell COMMENTS, not the notes column.
TARGET = ["mii_module", "mii_profile", "fhir_element", "transform", "status"]
_WB = {}

def _workbook(sheet_id):
    if sheet_id not in _WB:
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 genomDE2FHIR-sync/1.0"})
        data = urllib.request.urlopen(req, timeout=40).read()
        if data[:2] != b"PK":
            raise RuntimeError("got non-xlsx — is the sheet link-shared (anyone-with-link viewer)?")
        _WB[sheet_id] = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    return _WB[sheet_id]

def fetch_tab(sheet_id, tab):
    ws = _workbook(sheet_id)[tab]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hdr = [str(c) if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = {hdr[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(hdr)}
        if d.get("path"):
            out.append(d)
    return out

def main():
    args = sys.argv[1:]
    dry = "--dry-run" in args
    sid = SHEET_ID
    if "--id" in args:
        sid = args[args.index("--id") + 1]
    total_changed = 0
    for tab, fn in TABS.items():
        sheet_rows = fetch_tab(sid, tab)
        if not sheet_rows or "path" not in sheet_rows[0]:
            print(f"!! {tab}: no 'path' column — skipped"); continue
        sheet = {r["path"]: r for r in sheet_rows}
        local = list(csv.DictReader(open(ROOT / "mapping" / fn)))
        fields = list(local[0].keys())
        changed, diffs = 0, []
        local_paths = set()
        for r in local:
            local_paths.add(r["path"])
            s = sheet.get(r["path"])
            if not s:
                continue
            for col in TARGET:
                if col in s and (s[col] or "") != (r[col] or ""):
                    diffs.append(f"   {r['path']} | {col}: {r[col]!r} -> {s[col]!r}")
                    r[col] = s[col]; changed += 1
        only_sheet = [p for p in sheet if p not in local_paths]
        only_local = [p for p in local_paths if p not in sheet]
        print(f"== {tab} -> {fn}: {changed} cell change(s)" + (" [DRY-RUN]" if dry else ""))
        for d in diffs[:40]:
            print(d)
        if len(diffs) > 40:
            print(f"   ... +{len(diffs)-40} more")
        if only_sheet:
            print(f"   ⚠ {len(only_sheet)} path(s) in sheet not in CSV (ignored): {only_sheet[:3]}")
        if only_local:
            print(f"   ⚠ {len(only_local)} path(s) in CSV not in sheet (left as-is): {only_local[:3]}")
        if changed and not dry:
            with open(ROOT / "mapping" / fn, "w", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(local)
        total_changed += changed
    print(f"\nTotal: {total_changed} cell change(s)" + (" (dry-run, nothing written)" if dry else " written"))
    if total_changed and not dry:
        print("Review with `git diff mapping/`, then commit.")

if __name__ == "__main__":
    main()
